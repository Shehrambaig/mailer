import openpyxl
from datetime import datetime
from app.extensions import db
from app.models import Lead
from app.ingest.normalizer import (
    normalize_street, normalize_state, normalize_zip,
    safe_int, safe_float,
)


class ExcelLoader:
    def load_file(self, filepath):
        """Load all sheets from an Excel file. Returns total new leads inserted."""
        wb = openpyxl.load_workbook(filepath, read_only=True)
        total = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if headers and headers[0] == "Street":
                total += self._load_mailing_sheet(ws, headers)
        wb.close()
        return total

    def _load_mailing_sheet(self, ws, headers):
        """Load the simple mailing data sheet (8 columns)."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            street = normalize_street(data.get("Street"))
            city = (data.get("City") or "").strip().upper()
            state = normalize_state(data.get("State"))
            zip_code = normalize_zip(data.get("Zip Code"))

            if not street or not city or not state or not zip_code:
                continue

            auction_date = data.get("Auction Date")
            if isinstance(auction_date, str):
                try:
                    auction_date = datetime.strptime(auction_date, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    auction_date = None

            lead = Lead(
                source="excel_mailing",
                street=street,
                city=city,
                state=state,
                zip_code=zip_code,
                auction_date=auction_date,
                estimated_value=safe_int(data.get("Estimated Value")),
                offer_value=safe_int(data.get("Offer Value")),
                campaign_phone=data.get("Campaign Phone"),
            )
            db.session.merge(_upsert_lead(lead))
            count += 1

        db.session.commit()
        return count


def _upsert_lead(new_lead):
    """Find existing lead by address or return new one."""
    existing = Lead.query.filter_by(
        street=new_lead.street,
        city=new_lead.city,
        state=new_lead.state,
        zip_code=new_lead.zip_code,
    ).first()

    if existing:
        # Update with richer data if available
        for field in [
            "source_id", "source_url", "county", "property_type",
            "bedrooms", "bathrooms", "square_footage", "lot_size_acres",
            "year_built", "latitude", "longitude", "estimated_value",
            "offer_value", "starting_bid", "reserve_price", "auction_date",
            "auction_end_date", "auction_type", "asset_type", "product_type",
            "occupancy_status", "listing_status", "servicer_name",
            "investor_name", "attorney_name", "campaign_phone",
        ]:
            new_val = getattr(new_lead, field)
            if new_val is not None:
                setattr(existing, field, new_val)
        if new_lead.address_verified and not existing.address_verified:
            existing.address_verified = True
            existing.dpv_match_code = new_lead.dpv_match_code
        return existing

    return new_lead


