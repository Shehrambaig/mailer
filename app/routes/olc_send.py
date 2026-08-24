"""OpenLetterConnect send page + API.

Surfaced routes:

  GET  /send                          the compose/review/send UI
  GET  /api/olc/meta                  products + templates from OLC (or configured:false)
  POST /api/olc/recipients/parse      xlsx/csv upload -> validated rows + suppression flags
  GET  /api/olc/queue                 queued mailer rows (mailed_at IS NULL) -> same shape
  POST /api/olc/orders                place an order via OLC, log to olc_orders/items
  GET  /api/olc/orders                order history from Neon
  GET  /api/olc/orders/<id>/items     recipients for one order
  POST /api/olc/orders/<id>/refresh   poll OLC for current status / delivery info

Suppression ("have we already mailed this address?") is checked against three
send-history sources, each with its own address fingerprint:

  1. olc_order_items.addr_key       — orders placed through this page
  2. mailer table (mailed rows)     — send history; rows with mailed_at NULL
                                      are the send QUEUE and never suppress
  3. mailing_export_items.address_norm — mail-house xlsx exports

addr_key is the Foreclosure-Project canonical form so it agrees with the
xlsx dedupe pipeline:  NORM(street)|NORM(city)|NORM(state)|zip5
where NORM uppercases, maps non-[A-Z0-9 ] to space, collapses whitespace.
"""
from __future__ import annotations

import base64
import io
import json
import os
import re
from datetime import datetime, timezone

import psycopg2
import psycopg2.extras
from flask import Blueprint, Response, jsonify, render_template, request

from app.api.olc import OLCApiError, OLCClient, OLCNotConfigured, _unwrap

olc_bp = Blueprint("olc", __name__)

MAX_UPLOAD_ROWS = 20000


def _conn():
    dsn = (os.getenv("NEON_DB")
           or os.getenv("PROBATE_DATABASE_URL")
           or "postgresql://localhost:5432/probate")
    return psycopg2.connect(dsn, connect_timeout=10)


# ── Address fingerprints ─────────────────────────────────────────────────────

def norm_text(v) -> str:
    """Mirror Foreclosure-Project/mailer/dedup_mailer.py norm_text exactly."""
    if v is None:
        return ""
    s = str(v).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_zip(v) -> str:
    """Mirror dedup_mailer.py norm_zip: strip trailing .0, digits only, first 5."""
    if v is None:
        return ""
    s = re.sub(r"\.0$", "", str(v).strip())
    return re.sub(r"[^0-9]", "", s)[:5]


# USPS CASS expands "#3A" into "APT 3A", "SPC 16", "LOT 91" and so on, so the
# address OLC echoes back in a webhook rarely keys the same as the one we sent.
# Dropping the designator makes both sides agree.
UNIT_WORDS = {
    "APT", "APARTMENT", "STE", "SUITE", "UNIT", "SPC", "SPACE", "LOT",
    "RM", "ROOM", "FL", "FLOOR", "BLDG", "BUILDING", "TRLR", "TRAILER",
    "DEPT", "DEPARTMENT", "PMB", "BOX",
}


def loose_addr_key(street, city, state, zc) -> str:
    """addr_key with secondary-unit designators removed from the street."""
    words = [w for w in norm_text(street).split() if w not in UNIT_WORDS]
    return "|".join([" ".join(words), norm_text(city), norm_text(state), norm_zip(zc)])


def addr_key(street, city, state, zc) -> str:
    return "|".join([norm_text(street), norm_text(city), norm_text(state), norm_zip(zc)])


def export_addr_norm(street, zc) -> str | None:
    """Mirror mailing_export.ADDR_NORM_SQL: alnum-only street | LPAD(zip,5,'0')."""
    z = re.sub(r"\D", "", str(zc or ""))
    if not z:
        return None
    return re.sub(r"[^A-Za-z0-9]+", "", str(street or "")).upper() + "|" + z[:5].rjust(5, "0")


# ── Upload parsing ───────────────────────────────────────────────────────────

# canonical field -> accepted header spellings (headers normalized to alnum-lower)
HEADER_ALIASES = {
    "first_name":      {"firstname", "first", "ownerfirstname"},
    "last_name":       {"lastname", "last", "ownerlastname"},
    "street":          {"street", "address", "address1", "streetaddress", "mailingaddress", "propertyaddress"},
    "city":            {"city", "mailingcity", "propertycity"},
    "state":           {"state", "st", "mailingstate", "propertystate"},
    "zip_code":        {"zip", "zipcode", "zip5", "postalcode", "mailingzip", "propertyzip"},
    "auction_date":    {"auctiondate", "auction"},
    "estimated_value": {"estimatedvalue", "estvalue", "estimated"},
    "offer_value":     {"offervalue", "offer"},
    "campaign_phone":  {"campaignphone", "phone", "phonenumber"},
}


def _norm_header(h) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def _map_headers(header_row) -> dict[int, str]:
    """column index -> canonical field name (unknown columns are ignored)."""
    mapping = {}
    claimed = set()
    for idx, raw in enumerate(header_row):
        h = _norm_header(raw)
        if not h:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if field not in claimed and h in aliases:
                mapping[idx] = field
                claimed.add(field)
                break
    return mapping


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _parse_money(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$,\s]", "", str(v))
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if hasattr(v, "isoformat"):  # date
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _read_upload(file_storage) -> list[list]:
    """Return raw rows (first row = header) from an xlsx or csv upload."""
    filename = (file_storage.filename or "").lower()
    blob = file_storage.read()
    if len(blob) > 30 * 1024 * 1024:
        raise ValueError("File larger than 30 MB")

    if filename.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > MAX_UPLOAD_ROWS + 1:
                raise ValueError(f"More than {MAX_UPLOAD_ROWS} rows")
        wb.close()
        return rows

    if filename.endswith((".csv", ".txt")):
        import csv
        text = blob.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if len(rows) > MAX_UPLOAD_ROWS + 1:
            raise ValueError(f"More than {MAX_UPLOAD_ROWS} rows")
        return rows

    raise ValueError("Unsupported file type — upload .xlsx or .csv")


# SQL twin of norm_text(): uppercase, non-[A-Z0-9 ] -> space, collapse, trim.
def _norm_text_sql(col: str) -> str:
    return ("TRIM(REGEXP_REPLACE(REGEXP_REPLACE(UPPER(COALESCE(" + col + ", '')), "
            "'[^A-Z0-9 ]', ' ', 'g'), '\\s+', ' ', 'g'))")


# SQL twin of norm_zip(): strip trailing .0, digits only, first five.
def _norm_zip_sql(col: str) -> str:
    return ("LEFT(REGEXP_REPLACE(REGEXP_REPLACE(COALESCE(" + col + ", ''), "
            "'\\.0$', ''), '[^0-9]', '', 'g'), 5)")


MAILER_KEY_SQL = " || '|' || ".join([
    _norm_text_sql("street"), _norm_text_sql("city"),
    _norm_text_sql("state"), _norm_zip_sql("zip_code"),
])


def _suppression_sets(keys: list[str], export_norms: list[str]):
    """Look up which uploaded addresses were already mailed, per source.

    Each source is queried independently and a missing table just means
    "no history there" — the page must not 500 because one send-history
    table hasn't been created on this DB yet.
    """
    sent_olc, sent_mailer, exported = set(), set(), set()
    if not keys:
        return sent_olc, sent_mailer, exported

    conn = _conn()
    try:
        def _lookup(sql, params):
            try:
                with conn.cursor() as cur:
                    cur.execute(sql, params)
                    return {r[0] for r in cur.fetchall()}
            except psycopg2.errors.UndefinedTable:
                conn.rollback()
                return set()

        sent_olc = _lookup(
            "SELECT DISTINCT addr_key FROM olc_order_items WHERE addr_key = ANY(%s)",
            (keys,),
        )
        # mailer stores raw address parts (110k+ rows) — compute the key in
        # SQL so only the matches cross the wire. Only mailed rows count as
        # history; mailed_at NULL rows are the queue awaiting send.
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f"SELECT DISTINCT k FROM (SELECT {MAILER_KEY_SQL} AS k FROM mailer "
                    "WHERE mailed_at IS NOT NULL) m WHERE m.k = ANY(%s)",
                    (keys,),
                )
                sent_mailer = {r[0] for r in cur.fetchall()}
        except psycopg2.errors.UndefinedColumn:
            # pre-migration DB: no queue column means every row is history
            conn.rollback()
            sent_mailer = _lookup(
                f"SELECT DISTINCT k FROM (SELECT {MAILER_KEY_SQL} AS k FROM mailer) m "
                "WHERE m.k = ANY(%s)",
                (keys,),
            )
        except psycopg2.errors.UndefinedTable:
            conn.rollback()
            sent_mailer = set()
        if export_norms:
            exported = _lookup(
                "SELECT DISTINCT address_norm FROM mailing_export_items "
                "WHERE address_norm = ANY(%s)",
                (export_norms,),
            )
    finally:
        conn.close()
    return sent_olc, sent_mailer, exported


def _review_rows(rows: list[dict], *, history_blocks: bool = True) -> dict:
    """Validate, fingerprint, and suppression-flag canonical row dicts in
    place; returns the summary. With history_blocks=False (queue rows — the
    pipeline already applied the weekly dedupe rule before queueing), the
    send-history flags are shown as warnings but don't deselect the row."""
    seen_keys = set()
    for rec in rows:
        problems = []
        if not rec.get("street"):
            problems.append("missing street")
        if not rec.get("city"):
            problems.append("missing city")
        if not (rec.get("state") and len(rec["state"]) == 2):
            problems.append("missing/bad state")
        if not (rec.get("zip_code") and len(rec["zip_code"]) == 5):
            problems.append("missing/bad zip")

        key = addr_key(rec.get("street"), rec.get("city"),
                       rec.get("state"), rec.get("zip_code"))
        rec["addr_key"] = key
        rec["invalid"] = problems
        rec["dup_in_file"] = key in seen_keys and not problems
        if not problems:
            seen_keys.add(key)

    keys = sorted(seen_keys)
    export_norms = sorted({
        n for r in rows
        if not r["invalid"] and (n := export_addr_norm(r["street"], r["zip_code"]))
    })
    sent_olc, sent_mailer, exported = _suppression_sets(keys, export_norms)

    n_selected = 0
    for r in rows:
        r["sent_olc"] = r["addr_key"] in sent_olc
        r["sent_mailer"] = r["addr_key"] in sent_mailer
        r["exported"] = (export_addr_norm(r["street"], r["zip_code"]) or "") in exported
        r["no_name"] = not (r.get("first_name") or r.get("last_name"))
        # warn-only: the check template prints the offer — a missing value
        # means a blank amount box on the letter
        r["no_offer"] = r.get("offer_value") in (None, "")
        blocked = bool(r["invalid"]) or r["dup_in_file"]
        if history_blocks:
            blocked = blocked or r["sent_olc"] or r["sent_mailer"] or r["exported"]
        r["selected"] = not blocked
        n_selected += r["selected"]

    return {
        "total": len(rows),
        "invalid": sum(1 for r in rows if r["invalid"]),
        "dup_in_file": sum(1 for r in rows if r["dup_in_file"]),
        "sent_olc": sum(1 for r in rows if r["sent_olc"]),
        "sent_mailer": sum(1 for r in rows if r["sent_mailer"]),
        "exported": sum(1 for r in rows if r["exported"]),
        "no_name": sum(1 for r in rows if r["no_name"]),
        "no_offer": sum(1 for r in rows if r["no_offer"]),
        "selected": n_selected,
    }


@olc_bp.route("/api/olc/recipients/parse", methods=["POST"])
def parse_recipients():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded (field name must be 'file')"}), 400
    try:
        raw = _read_upload(request.files["file"])
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if len(raw) < 2:
        return jsonify({"error": "File has no data rows"}), 400

    col_map = _map_headers(raw[0])
    mapped_fields = set(col_map.values())
    missing = [f for f in ("street", "city", "state", "zip_code") if f not in mapped_fields]
    if missing:
        return jsonify({
            "error": f"Could not find required column(s): {', '.join(missing)}. "
                     f"Headers seen: {[str(h) for h in raw[0] if h]}",
        }), 400

    rows = []
    for raw_row in raw[1:]:
        if raw_row is None or all(c in (None, "") for c in raw_row):
            continue
        rec = {f: None for f in HEADER_ALIASES}
        for idx, field in col_map.items():
            v = raw_row[idx] if idx < len(raw_row) else None
            if field == "auction_date":
                rec[field] = _parse_date(v)
            elif field in ("estimated_value", "offer_value"):
                rec[field] = _parse_money(v)
            elif field == "zip_code":
                rec[field] = norm_zip(v)
            elif field == "state":
                rec[field] = _cell_str(v).upper() or None
            else:
                rec[field] = _cell_str(v) or None
        rows.append(rec)

    summary = _review_rows(rows, history_blocks=True)
    return jsonify({"rows": rows, "summary": summary, "source": "file"})


@olc_bp.route("/api/olc/queue", methods=["GET"])
def load_queue():
    """Queued mailer rows (mailed_at IS NULL) shaped exactly like a parsed
    upload, so the /send review UI treats both sources the same. The pipeline
    lands rows fully sendable (names, offer_value, campaign_phone computed)
    and already applied the weekly dedupe rule — history flags warn only."""
    conn = _conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            try:
                cur.execute("""
                    SELECT id, first_name, last_name, street, city, state, zip_code,
                           auction_date, estimated_value, offer_value, campaign_phone,
                           source_file, batch, loaded_at,
                           COUNT(*) OVER () AS queued_total
                    FROM mailer
                    WHERE mailed_at IS NULL
                    ORDER BY id
                    LIMIT %s
                """, (MAX_UPLOAD_ROWS,))
                db_rows = [dict(r) for r in cur.fetchall()]
            except psycopg2.errors.UndefinedColumn:
                conn.rollback()
                return jsonify({"error": "mailer has no mailed_at column yet — run "
                                         "scripts/add_mailer_queue_columns.py"}), 500
    finally:
        conn.close()

    queued_total = db_rows[0]["queued_total"] if db_rows else 0
    rows = []
    for d in db_rows:
        rows.append({
            "mailer_id": d["id"],
            "first_name": _cell_str(d["first_name"]) or None,
            "last_name": _cell_str(d["last_name"]) or None,
            "street": _cell_str(d["street"]) or None,
            "city": _cell_str(d["city"]) or None,
            "state": (_cell_str(d["state"]).upper() or None),
            "zip_code": norm_zip(d["zip_code"]) or None,
            "auction_date": d["auction_date"].isoformat() if d["auction_date"] else None,
            "estimated_value": float(d["estimated_value"]) if d["estimated_value"] is not None else None,
            "offer_value": float(d["offer_value"]) if d["offer_value"] is not None else None,
            "campaign_phone": _cell_str(d["campaign_phone"]) or None,
            "batch": d["batch"] or d["source_file"],
        })

    summary = _review_rows(rows, history_blocks=False)
    return jsonify({"rows": rows, "summary": summary, "source": "queue",
                    "queued_total": queued_total,
                    "truncated": queued_total > len(rows)})


# ── Meta (products + templates for the compose form) ────────────────────────

@olc_bp.route("/api/olc/meta", methods=["GET"])
def meta():
    try:
        client = OLCClient.from_env()
    except OLCNotConfigured as e:
        return jsonify({"configured": False, "message": str(e),
                        "products": [], "templates": []})
    try:
        products = client.get_products()
        templates = client.get_templates()
    except OLCApiError as e:
        return jsonify({"configured": True, "live": client.is_live,
                        "error": f"{e.status_code}: {e.message}",
                        "products": [], "templates": []})
    return jsonify({"configured": True, "live": client.is_live,
                    "base_url": client.base_url,
                    "products": products, "templates": templates})


# ── Place order ──────────────────────────────────────────────────────────────

_ONES = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight",
         "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
         "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy",
         "Eighty", "Ninety"]


def _words_under_1000(n: int) -> str:
    parts = []
    if n >= 100:
        parts.append(f"{_ONES[n // 100]} Hundred")
        n %= 100
    if n >= 20:
        word = _TENS[n // 10]
        if n % 10:
            word += f"-{_ONES[n % 10]}"
        parts.append(word)
    elif n:
        parts.append(_ONES[n])
    return " ".join(parts)


def dollars_in_words(amount) -> str:
    """175000 -> 'One Hundred Seventy-Five Thousand Dollars' (check amount line)."""
    n = int(round(float(amount)))
    if n <= 0:
        return ""
    if n >= 1_000_000_000:
        return ""  # out of scope for a mailer offer — leave blank rather than wrong
    parts = []
    for scale, label in ((1_000_000, "Million"), (1_000, "Thousand"), (1, "")):
        if n >= scale:
            chunk = _words_under_1000(n // scale)
            parts.append(f"{chunk} {label}".strip())
            n %= scale
    return " ".join(parts) + " Dollars"


def _olc_contact(rec: dict) -> dict:
    """Uploaded row -> OLC contact object. Mailing address doubles as the
    property address (we mail the foreclosure property itself).

    Check-window workaround (2026-08-07): the snap-pack window prints
    "FIRST LAST" then companyName then the address, so the name fields are
    hijacked to make the payee block read "PAY TO THE ORDER OF" above the
    recipient's real name (carried in companyName). Templates must therefore
    greet with {{CF.GREETING_NAME}} (custom field 1637) — {{C.FIRST_NAME}}
    would render "PAY TO THE"."""
    full_name = " ".join(
        s for s in ((rec.get("first_name") or "").strip(),
                    (rec.get("last_name") or "").strip()) if s
    ) or "Current Resident"
    c = {
        "firstName": "PAY TO THE",
        "lastName": "ORDER OF:",
        "companyName": full_name,
        "address1": rec["street"],
        "city": rec["city"],
        "state": rec["state"],
        "zip": rec["zip_code"],
        "propertyAddress": rec["street"],
        "propertyCity": rec["city"],
        "propertyState": rec["state"],
        "propertyZip": rec["zip_code"],
    }
    if rec.get("campaign_phone"):
        c["campaign_phone"] = rec["campaign_phone"]
    # Custom-field values match by the field's display name — sent both flat
    # (what view-proof reads) and nested under customFields (the documented
    # contact shape) so both merge paths work.
    # the template's salutation line is just "{{CF.GREETING_NAME}}," — the
    # value carries "Dear <name>" for people; entity rows would pass the bare
    # entity name (no Dear) when the queue starts carrying them
    first = (rec.get("first_name") or "").strip().title()
    cf = {"Greeting Name": f"Dear {first}" if first else "Dear Neighbor"}
    if rec.get("offer_value") not in (None, ""):
        try:
            cf["Offer Amount"] = f"${float(rec['offer_value']):,.0f}"
            words = dollars_in_words(rec["offer_value"])
            # the 11687-family check art prints its own "Dollars" label at
            # the end of the payline — don't double it
            if words.endswith(" Dollars"):
                words = words[:-len(" Dollars")]
            if words:
                cf["Offer Amount Words"] = words
        except (TypeError, ValueError):
            pass
    elif not (str(rec.get("first_name") or "").strip()
              and str(rec.get("last_name") or "").strip()):
        # "Call for Price" rule (user/Jessie 2026-08-06, producer handoff
        # §A12): rows without BOTH names carry offer_value NULL — the queue
        # producer nulls them deliberately. Numeric auto-formatting can't
        # fire, so map the literal text; never send a blank check amount.
        cf["Offer Amount"] = "Call for Price"
        cf["Offer Amount Words"] = "Call for Price"
    c.update(cf)
    c["customFields"] = cf
    return c


def _return_address() -> dict | None:
    """Sender block from env — feeds the template's SPF.* merge fields.
    Returns None (OLC falls back to account settings) unless the full
    address is configured."""
    fields = {
        "firstName": os.getenv("OLC_RETURN_FIRST_NAME"),
        "lastName": os.getenv("OLC_RETURN_LAST_NAME"),
        "companyName": os.getenv("OLC_RETURN_COMPANY"),
        "address1": os.getenv("OLC_RETURN_ADDRESS1"),
        "address2": os.getenv("OLC_RETURN_ADDRESS2"),
        "city": os.getenv("OLC_RETURN_CITY"),
        "state": os.getenv("OLC_RETURN_STATE"),
        "zip": os.getenv("OLC_RETURN_ZIP"),
        "phoneNo": os.getenv("OLC_RETURN_PHONE"),
        "email": os.getenv("OLC_RETURN_EMAIL"),
    }
    fields = {k: v for k, v in fields.items() if v}
    return fields if {"address1", "city", "state", "zip"} <= fields.keys() else None


@olc_bp.route("/api/olc/orders", methods=["POST"])
def place_order():
    body = request.get_json(silent=True) or {}
    contacts = body.get("contacts") or []
    product_id = body.get("productId")
    template_id = body.get("templateId")
    name = (body.get("name") or "").strip()

    if not contacts:
        return jsonify({"error": "No recipients selected"}), 400
    if not product_id or not template_id:
        return jsonify({"error": "productId and templateId are required"}), 400
    if not name:
        name = f"send-{datetime.now(timezone.utc).strftime('%Y-%m-%d-%H%M')}"

    bad = [c for c in contacts
           if not (c.get("street") and c.get("city") and c.get("state") and c.get("zip_code"))]
    if bad:
        return jsonify({"error": f"{len(bad)} recipient(s) missing address fields"}), 400

    try:
        client = OLCClient.from_env()
    except OLCNotConfigured as e:
        return jsonify({"error": str(e), "setup": True}), 400

    payload = {
        "name": name,
        "productId": int(product_id),
        "templateId": int(template_id),
        "contacts": [_olc_contact(c) for c in contacts],
    }
    return_address = _return_address()
    if return_address:
        payload["returnAddress"] = return_address

    try:
        resp = client.create_order(payload)
    except OLCApiError as e:
        return jsonify({"error": f"OLC rejected the order ({e.status_code}): {e.message}",
                        "data": e.data}), 502

    data = _unwrap(resp) or {}
    now = datetime.now(timezone.utc)

    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO olc_orders
                    (olc_order_id, name, product_id, product_name,
                     template_id, template_name, status, cost, is_live_mode,
                     recipient_count, source_note, response_json)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id
            """, (
                str(data.get("id")) if data.get("id") is not None else None,
                name, int(product_id), body.get("productName"),
                int(template_id), body.get("templateName"),
                data.get("status"), data.get("cost"),
                data.get("isLiveMode", client.is_live),
                len(contacts), body.get("sourceNote"), json.dumps(resp, default=str),
            ))
            local_id = cur.fetchone()[0]

            psycopg2.extras.execute_values(cur, """
                INSERT INTO olc_order_items
                    (order_id, first_name, last_name, street, city, state, zip_code,
                     campaign_phone, auction_date, estimated_value, offer_value,
                     addr_key, status, updated_at)
                VALUES %s
            """, [(
                local_id, c.get("first_name"), c.get("last_name"),
                c.get("street"), c.get("city"), c.get("state"), c.get("zip_code"),
                c.get("campaign_phone"), c.get("auction_date"),
                c.get("estimated_value"), c.get("offer_value"),
                addr_key(c.get("street"), c.get("city"), c.get("state"), c.get("zip_code")),
                data.get("status"), now,
            ) for c in contacts])

            # queue-sourced rows carry their mailer.id — stamp them sent so
            # they leave the queue and become suppression history
            stamped = 0
            mailer_ids = [int(c["mailer_id"]) for c in contacts if c.get("mailer_id")]
            if mailer_ids:
                cur.execute("""
                    UPDATE mailer SET mailed_at = %s, olc_order_id = %s
                    WHERE id = ANY(%s) AND mailed_at IS NULL
                """, (now, local_id, mailer_ids))
                stamped = cur.rowcount
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True, "order": {
        "id": local_id,
        "olc_order_id": data.get("id"),
        "name": name,
        "status": data.get("status"),
        "cost": data.get("cost"),
        "is_live_mode": data.get("isLiveMode", client.is_live),
        "recipient_count": len(contacts),
        "queue_rows_stamped": stamped,
    }})


@olc_bp.route("/api/olc/proof", methods=["POST"])
def proof():
    """Render one recipient through a template via OLC's view-proof — returns
    the PDF inline so the browser can open it. Nothing is ordered or mailed."""
    body = request.get_json(silent=True) or {}
    template_id = body.get("templateId")
    rec = body.get("contact") or {}
    if not template_id:
        return jsonify({"error": "templateId is required"}), 400
    if not (rec.get("street") and rec.get("city") and rec.get("state") and rec.get("zip_code")):
        return jsonify({"error": "Contact needs street/city/state/zip"}), 400

    try:
        client = OLCClient.from_env()
    except OLCNotConfigured as e:
        return jsonify({"error": str(e), "setup": True}), 400

    payload = {"templateId": int(template_id), "contact": _olc_contact(rec)}
    return_address = _return_address()
    if return_address:
        payload["returnContact"] = return_address

    try:
        resp = client.view_proof(payload)
    except OLCApiError as e:
        return jsonify({"error": f"OLC proof failed ({e.status_code}): {e.message}"}), 502

    b64 = (resp.get("data") or {}).get("base64") if isinstance(resp, dict) else None
    if not b64:
        return jsonify({"error": "OLC returned no PDF"}), 502
    return Response(
        base64.b64decode(b64),
        mimetype="application/pdf",
        headers={"Content-Disposition": 'inline; filename="olc-proof.pdf"'},
    )


# ── Order history / tracking ────────────────────────────────────────────────

@olc_bp.route("/api/olc/orders", methods=["GET"])
def list_orders():
    conn = _conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT o.id, o.created_at, o.olc_order_id, o.name,
                       o.product_id, o.product_name, o.template_id, o.template_name,
                       o.status, o.cost, o.is_live_mode, o.recipient_count,
                       o.source_note, o.last_polled_at,
                       COUNT(i.id)                                   AS item_count,
                       COUNT(i.id) FILTER (WHERE i.is_delivered)     AS delivered_count
                FROM olc_orders o
                LEFT JOIN olc_order_items i ON i.order_id = o.id
                GROUP BY o.id
                ORDER BY o.created_at DESC
                LIMIT 200
            """)
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    for r in rows:
        for k in ("created_at", "last_polled_at"):
            if r.get(k):
                r[k] = r[k].isoformat()
        if r.get("cost") is not None:
            r["cost"] = float(r["cost"])
    return jsonify({"orders": rows})


@olc_bp.route("/api/olc/orders/<int:order_id>/items", methods=["GET"])
def order_items(order_id: int):
    conn = _conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT first_name, last_name, street, city, state, zip_code,
                       campaign_phone, auction_date, estimated_value, offer_value,
                       status, is_delivered, tracking_code, updated_at
                FROM olc_order_items
                WHERE order_id = %s
                ORDER BY id
            """, (order_id,))
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    for r in rows:
        for k in ("auction_date", "updated_at"):
            if r.get(k):
                r[k] = r[k].isoformat()
        for k in ("estimated_value", "offer_value"):
            if r.get(k) is not None:
                r[k] = float(r[k])
    return jsonify({"items": rows})


@olc_bp.route("/api/olc/orders/<int:order_id>/refresh", methods=["POST"])
def refresh_order(order_id: int):
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT olc_order_id FROM olc_orders WHERE id = %s", (order_id,))
            row = cur.fetchone()
        if not row:
            return jsonify({"error": "Unknown order"}), 404
        olc_order_id = row[0]
        if not olc_order_id:
            return jsonify({"error": "Order has no OLC id recorded"}), 400

        try:
            client = OLCClient.from_env()
        except OLCNotConfigured as e:
            return jsonify({"error": str(e), "setup": True}), 400

        data = None
        try:
            data = _unwrap(client.get_order(olc_order_id))
        except OLCApiError:
            # Some deployments only expose the list endpoint — fall back to search.
            try:
                listing = _unwrap(client.get_orders(search=str(olc_order_id)))
                rows = listing.get("rows") if isinstance(listing, dict) else listing
                for cand in rows or []:
                    if str(cand.get("id")) == str(olc_order_id):
                        data = cand
                        break
            except OLCApiError as e:
                return jsonify({"error": f"OLC error ({e.status_code}): {e.message}"}), 502
        if not isinstance(data, dict):
            return jsonify({"error": "Order not found on OLC side"}), 404

        now = datetime.now(timezone.utc)
        status = data.get("status") or data.get("orderStatus")

        with conn.cursor() as cur:
            cur.execute("""
                UPDATE olc_orders
                SET status = COALESCE(%s, status), last_polled_at = %s
                WHERE id = %s
            """, (status, now, order_id))

            updated_items = 0
            for contact in data.get("contacts") or []:
                k = addr_key(
                    contact.get("address1") or contact.get("address"),
                    contact.get("city"), contact.get("state"),
                    contact.get("zip") or contact.get("zipCode"),
                )
                cur.execute("""
                    UPDATE olc_order_items
                    SET status = COALESCE(%s, status),
                        is_delivered = COALESCE(%s, is_delivered),
                        tracking_code = COALESCE(%s, tracking_code),
                        updated_at = %s
                    WHERE order_id = %s AND addr_key = %s
                """, (contact.get("status"), contact.get("isDelivered"),
                      contact.get("trackingCode"), now, order_id, k))
                updated_items += cur.rowcount
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True, "status": status, "items_updated": updated_items,
                    "polled_at": now.isoformat()})


# ── Page ─────────────────────────────────────────────────────────────────────

@olc_bp.route("/send")
def send_page():
    return render_template("send.html")
