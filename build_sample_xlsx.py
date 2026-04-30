import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path("/Users/shehrambaig/PycharmProjects/mailer/Mailer_Data_set")
OUT = Path("/Users/shehrambaig/PycharmProjects/mailer/Mailer_Data_set_sample.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Samples"

label_font = Font(bold=True, color="FFFFFF")
label_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")

row = 1
max_cols = 0

files = sorted(p for p in ROOT.rglob("*.csv") if p.is_file())

for path in files:
    rel = path.relative_to(ROOT)
    parts = rel.parts
    source = parts[0]
    subfolder = "/".join(parts[1:-1]) if len(parts) > 2 else ""
    file_name = path.name

    label = f"Source: {source}"
    if subfolder:
        label += f"  |  Folder: {subfolder}"
    label += f"  |  File: {file_name}"

    ws.cell(row=row, column=1, value=label).font = label_font
    ws.cell(row=row, column=1).fill = label_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    row += 1

    raw = path.read_bytes()
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        encoding = "utf-16"
    elif raw.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    else:
        encoding = "utf-8"

    try:
        text = raw.decode(encoding, errors="replace")
    except Exception:
        text = raw.decode("latin-1", errors="replace")

    sample = text[:4096]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","

    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    header = next(reader, [])
    data_rows = []
    for _ in range(3):
        try:
            data_rows.append(next(reader))
        except StopIteration:
            break

    def clean(v):
        if v is None:
            return v
        return "".join(ch for ch in str(v) if ch == "\t" or ch == "\n" or ord(ch) >= 32 or ch == "\r")

    header = [clean(v) for v in header]
    data_rows = [[clean(v) for v in r] for r in data_rows]

    for c, val in enumerate(header, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = header_font
        cell.fill = header_fill
    max_cols = max(max_cols, len(header))
    row += 1

    for data in data_rows:
        for c, val in enumerate(data, start=1):
            ws.cell(row=row, column=c, value=val)
        max_cols = max(max_cols, len(data))
        row += 1

    row += 1  # blank separator

# Merge label cells across the data width for readability
for r in range(1, row):
    cell = ws.cell(row=r, column=1)
    if cell.fill.fgColor and cell.fill.fgColor.rgb == "FF305496":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(max_cols, 1))

# Reasonable column widths
for c in range(1, max_cols + 1):
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 22

wb.save(OUT)
print(f"Wrote {OUT}")
